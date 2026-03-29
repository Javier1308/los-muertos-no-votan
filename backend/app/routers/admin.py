import io
from fastapi import APIRouter, Depends, HTTPException, Header, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func
from app.database import get_db
from app.models.report import Report
from app.schemas.report import AdminStatsResponse
from app.config import settings

router = APIRouter(prefix="/admin", tags=["admin"])


def verify_admin(x_api_key: str = Header(...)):
    if x_api_key != settings.ADMIN_API_KEY:
        raise HTTPException(status_code=403, detail="Acceso no autorizado")
    return x_api_key


@router.get("/stats", response_model=AdminStatsResponse)
async def estadisticas(
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_admin),
):
    total_result = await db.execute(select(func.count(Report.id)))
    total = total_result.scalar_one()

    parentesco_result = await db.execute(
        select(Report.parentesco, func.count(Report.id))
        .group_by(Report.parentesco)
        .order_by(func.count(Report.id).desc())
    )
    por_parentesco = {row[0]: row[1] for row in parentesco_result.fetchall()}

    mesa_result = await db.execute(
        select(Report.mesa_votacion, func.count(Report.id))
        .group_by(Report.mesa_votacion)
        .order_by(func.count(Report.id).desc())
        .limit(50)
    )
    por_mesa = {row[0]: row[1] for row in mesa_result.fetchall()}

    return AdminStatsResponse(
        total=total,
        por_parentesco=por_parentesco,
        por_mesa=por_mesa,
    )


@router.get("/reports/export/excel")
async def exportar_excel(
    parentesco: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    _: str = Depends(verify_admin),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    query = select(Report).order_by(Report.created_at.desc())
    if parentesco:
        query = query.where(Report.parentesco == parentesco)

    result = await db.execute(query)
    reports = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Denuncias"

    headers = [
        "ID", "DNI Denunciante", "DNI Denunciado",
        "Mesa de Votación", "Parentesco", "IP Registro", "Fecha Registro",
    ]

    header_fill = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, r in enumerate(reports, 2):
        created_at_naive = r.created_at.replace(tzinfo=None) if r.created_at.tzinfo else r.created_at
        ws.append([
            str(r.id),
            r.dni_denunciante,
            r.dni_denunciado,
            r.mesa_votacion,
            r.parentesco,
            str(r.ip_address) if r.ip_address else "",
            created_at_naive,
        ])
        ws.cell(row=row, column=7).number_format = "YYYY-MM-DD HH:MM:SS"

    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"denuncias{'_' + parentesco if parentesco else ''}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
